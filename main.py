import logging
import os
import sys
import traceback
from typing import Any, Dict

import pandas as pd
from playwright.sync_api import sync_playwright, Page
import tkinter as tk
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import updater
from unit_selector import select_units_and_login


current_version = 'v0.0.2'  # Versão atual do aplicativo

url_login_canaime = 'https://canaime.com.br/sgp2rr/login/login_principal.php'
url_reports = "https://canaime.com.br/sgp2rr/areas/unidades/Informes_LER.php?id_cad_preso="
url_certificate = "https://canaime.com.br/sgp2rr/areas/impressoes/UND_CertidaoCarceraria.php?id_cad_preso="
url_call = 'https://canaime.com.br/sgp2rr/areas/impressoes/UND_ChamadaFOTOS_todos2.php?id_und_prisional='
url_main = 'https://canaime.com.br/sgp2rr/areas/unidades/cadastro.php?id_cad_preso='
units = ('PAMC', 'CPBV', 'CPFBV', 'CPP', 'UPRRO')


class Logger:
    """
    Logger class to handle error logging to a file.

    Methods
    -------
    capture_error(error: Exception, page: Page = None) -> None
        Captures and logs the error with traceback and current URL.
    """

    @staticmethod
    def capture_error(error: Exception, page: Page = None) -> None:
        """
        Captures and logs the error with traceback and current URL.

        Parameters
        ----------
        error : Exception
            The exception that occurred.
        page : Page, optional
            The Playwright page where the error occurred (default is None).
        """
        logger = logging.getLogger()
        logger.setLevel(logging.ERROR)

        handler = logging.FileHandler('error_log.log', encoding='utf-8')
        handler.setLevel(logging.ERROR)

        formatter = logging.Formatter('%(asctime)s:%(levelname)s:%(message)s')
        handler.setFormatter(formatter)

        logger.addHandler(handler)

        error_message = f'Ocorreu um erro: {str(error)}'
        traceback_message = traceback.format_exc()
        current_url = page.url if page else 'URL não disponível'

        logger.error(f'{error_message}\nURL Atual: {current_url}\nTraceback:\n{traceback_message}')

        logger.removeHandler(handler)

        os.startfile('error_log.log')


class CanaimeLogin:
    """
    Handles login to the Canaimé website.

    Methods
    -------
    __init__(p: Any, headless: bool = True, login: str = '', password: str = '')
        Initializes the login handler with Playwright instance and settings.
    perform_login() -> Page
        Performs the login and returns the authenticated Playwright page.
    """

    def __init__(self, p: Any, headless: bool = True, login: str = '', password: str = ''):
        """
        Initializes the login handler with Playwright instance and settings.

        Parameters
        ----------
        p : Any
            Playwright instance.
        headless : bool, optional
            If True, run the browser in headless mode (default is True).
        login : str
            User login credential.
        password : str
            User password.
        """
        self.p = p
        self.headless = headless
        self.login = login
        self.password = password
        self.page = None

    def perform_login(self) -> Page:
        """
        Performs the login and returns the authenticated Playwright page.

        Returns
        -------
        Page
            Authenticated Playwright page.
        """
        os.system('cls' if os.name == 'nt' else 'clear')

        browser = self.p.chromium.launch(headless=self.headless)
        context = browser.new_context(java_script_enabled=False)

        # Block image loading
        context.set_extra_http_headers({"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"})
        context.route(
            "**/*",
            lambda route: route.abort() if route.request.resource_type == "image" else route.continue_()
        )

        self.page = context.new_page()
        self.page.goto(url_login_canaime, timeout=0)
        self.page.locator("input[name=\"usuario\"]").click()
        self.page.locator("input[name=\"usuario\"]").fill(self.login)
        self.page.locator("input[name=\"senha\"]").fill(self.password)
        self.page.locator("input[name=\"senha\"]").press("Enter")
        self.page.wait_for_timeout(5000)
        try:
            if self.page.locator('img').count() < 4:
                print('Usuário ou senha inválidos')
                sys.exit(1)
            else:
                print('Login efetuado com sucesso!')
        except Exception as e1:
            Logger.capture_error(e1, self.page)
            sys.exit(1)

        return self.page


class UnitProcessor:
    """
    Processes prison units and retrieves detailed information about inmates.

    Methods
    -------
    __init__(page: Page)
        Initializes the processor with the Playwright page.
    create_unit_list(unit: str) -> pd.DataFrame
        Creates a list of inmates for the specified prison unit.
    get_inmate_info(code: str) -> Dict[str, Any]
        Retrieves detailed information about an inmate.
    """

    def __init__(self, page: Page):
        """
        Initializes the processor with the Playwright page.

        Parameters
        ----------
        page : Page
            Playwright page instance.
        """
        self.page = page

    def create_unit_list(self, unit: str) -> pd.DataFrame:
        """
        Creates a list of inmates for the specified prison unit.

        Parameters
        ----------
        unit : str
            Prison unit code.

        Returns
        -------
        pd.DataFrame
            DataFrame with the list of inmates.
        """
        unit_list = pd.DataFrame(columns=['Ala', 'Cela', 'Código', 'Preso'])
        try:
            self.page.goto(url_call + unit, timeout=0)
            all_entries = self.page.locator('.titulobkSingCAPS')
            names = self.page.locator('.titulobkSingCAPS .titulo12bk')
            count = all_entries.count()
            for i in range(count):
                processed_entry = all_entries.nth(i).text_content().replace(" ", "").strip()
                [code, _, _, _, wing_cell] = processed_entry.split('\n')
                inmate = names.nth(i).text_content().strip()

                # Remove "ALA:" and separate by the last "/"
                wing_cell = wing_cell.replace("ALA:", "")
                split_index = wing_cell.rfind('/')
                wing = wing_cell[:split_index].strip()
                cell = wing_cell[split_index + 1:].strip()

                unit_list.loc[len(unit_list)] = [wing, cell, code[2:], inmate]
        except Exception as e2:
            Logger.capture_error(e2, self.page)
            sys.exit(1)

        return unit_list


def visualize_data(df: pd.DataFrame, message_box: tk.Text) -> None:
    """
    Generates a visual report of the number of cells and inmates per wing.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with inmate data organized by wing and cell.
    message_box : tk.Text
        Text widget for displaying messages.
    """
    report = df.groupby(['Ala', 'Cela']).size().reset_index(name='Qtd')
    wing_summary = report.groupby('Ala').agg({'Cela': 'nunique', 'Qtd': 'sum'}).reset_index()
    wing_summary = wing_summary.sort_values(by='Ala')  # Order by wing name

    message_box.insert(tk.END, "\nRelatório por Ala:\n")
    for index, row in wing_summary.iterrows():
        message_box.insert(tk.END,
                           f"Ala: {row['Ala']}, Qtd: {row['Cela']}, Total de Presos: {row['Qtd']}\n")
        cell_details = report[report['Ala'] == row['Ala']]
        for _, cell_row in cell_details.iterrows():
            message_box.insert(tk.END,
                               f"  Cela: {cell_row['Cela']}, Qtd: {cell_row['Qtd']}\n")


def create_excel_report(data: Dict[str, pd.DataFrame], filename: str) -> None:
    """
    Creates an Excel report with each unit as a separate sheet.

    Parameters
    ----------
    data : dict
        Dictionary where keys are unit names and values are DataFrames containing inmate data.
    filename : str
        Name of the Excel file to be created.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for unit, df in data.items():
        ws = wb.create_sheet(title=unit)

        # Group data by 'Ala' and process each group
        grouped = df.groupby('Ala')
        col_offset = 1

        for ala, group in grouped:
            # Sort celas by number (convert to int for sorting)
            cell_counts = group['Cela'].value_counts().sort_index(key=lambda x: x.astype(int), ascending=True)

            total_cells = cell_counts.size
            total_inmates = cell_counts.sum()
            average_inmates_per_cell = total_inmates / total_cells

            # Write Ala title in larger font, merge col1 and col2
            ala_title_cell = ws.cell(row=1, column=col_offset)
            ala_title_cell.value = f"Ala {ala}"
            ala_title_cell.font = Font(size=14, bold=True)
            ala_title_cell.alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset + 1)

            # Write number of cells and inmates
            ws.cell(row=2, column=col_offset).value = f"{total_cells} cela(s)"
            ws.cell(row=2, column=col_offset + 1).value = f"{total_inmates} preso(s)"

            # Write average inmates per cell, in italic, merge col1 and col2
            avg_inmates_cell = ws.cell(row=3, column=col_offset)
            avg_inmates_cell.value = f"{average_inmates_per_cell:.1f} preso(s) por cela"
            avg_inmates_cell.font = Font(size=10, italic=True)
            avg_inmates_cell.alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=3, start_column=col_offset, end_row=3, end_column=col_offset + 1)

            # Write header for cell details
            ws.cell(row=4, column=col_offset).value = "Cela"
            ws.cell(row=4, column=col_offset + 1).value = "Qtd"

            # Write data for each cell in the ala
            row_offset = 5
            for cell_name, inmate_count in cell_counts.items():
                ws.cell(row=row_offset, column=col_offset).value = cell_name
                ws.cell(row=row_offset, column=col_offset + 1).value = inmate_count
                row_offset += 1

            # Set column width and insert blank column between alas
            ws.column_dimensions[get_column_letter(col_offset)].width = 12
            ws.column_dimensions[get_column_letter(col_offset + 1)].width = 12
            ws.column_dimensions[get_column_letter(col_offset + 2)].width = 1
            col_offset += 3

    wb.save(filename)


def main(headless: bool = True) -> None:
    """
    Main function to handle the login and processing of inmate data from selected units.

    Parameters
    ----------
    headless : bool, optional
        If True, run the browser in headless mode (default is True).
    """
    # Step 1: Get user login and unit selection
    login, password, selected_units = select_units_and_login()  # Get user input

    with sync_playwright() as p:
        # Step 2: Perform login using the captured credentials
        login_handler = CanaimeLogin(p, headless=headless, login=login, password=password)
        page = login_handler.perform_login()  # Use the renamed method

        processor = UnitProcessor(page)

        # Step 3: Collect data for each selected prison unit
        all_units_data = {}
        for unit in selected_units:
            try:
                unit_data = processor.create_unit_list(unit)
                all_units_data[unit] = unit_data
                print(f"Unidade {unit} processada.")
            except Exception as e:
                Logger.capture_error(e, page)
                print(f"Falha ao processar a unidade {unit}, erro capturado.")

        # Step 4: Create Excel report with data from all units
        if all_units_data:
            create_excel_report(all_units_data, 'Presos por Ala.xlsx')
            print('Arquivo "Presos por Ala.xlsx" salvo.')

        print("Processo Compelto.")


if __name__ == '__main__':
    if updater.update_application(current_version):
        sys.exit(0)  # Exits the current application if an update is applied
    main()
