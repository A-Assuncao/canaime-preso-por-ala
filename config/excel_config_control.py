from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from datetime import datetime
import json
import os

# Define o diretório base relativo à localização do arquivo atual
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def load_units_config():
    # Constrói o caminho absoluto para o arquivo de configuração
    config_path = os.path.join(BASE_DIR, 'units_config.json')
    # Carregar configurações do arquivo JSON
    with open(config_path, 'r', encoding='utf-8') as file:
        config_data = json.load(file)
    return config_data


def calculate_shift(date):
    # Plantões: ALFA, BRAVO, CHARLIE, DELTA (se repetem a cada 4 dias)
    shifts = ["ALFA", "BRAVO", "CHARLIE", "DELTA"]
    start_date = datetime(2024, 1, 1)  # 1º de janeiro de 2024 foi CHARLIE
    days_since_start = (date - start_date).days
    shift_index = (days_since_start + 2) % 4  # Ajuste de índice conforme necessário
    return shifts[shift_index]


def generate_unit_control_sheet(workbook, unit_name):
    units_config = load_units_config()
    pamc_config = units_config[unit_name]["blocks"]

    # Cria a aba com o nome da unidade e "CONTROLE"
    sheet = workbook.create_sheet(title=f"{unit_name} CONTROLE")
    gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    purple_fill = PatternFill(start_color="D9D9F3", end_color="D9D9F3", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Adicionando bordas de A1 até B33
    for row in sheet['A1:B33']:
        for cell in row:
            cell.border = thin_border

    # Adicionar bordas para B35, B36, O13 e O14
    sheet['B35'].border = thin_border
    sheet['B36'].border = thin_border
    sheet['O13'].border = thin_border
    sheet['O14'].border = thin_border

    # Cabeçalho de Data e Plantão
    current_date = datetime.today()
    shift_name = calculate_shift(current_date)

    # Mesclar e preencher células para Data e Plantão
    sheet.merge_cells('A1:B1')
    sheet['A1'] = current_date.strftime("%d/%m/%Y")
    sheet['A1'].fill = purple_fill
    sheet['A1'].alignment = center_alignment
    sheet['A1'].font = bold_font
    sheet['A1'].border = thin_border

    sheet.merge_cells('A2:B2')
    sheet['A2'] = shift_name
    sheet['A2'].fill = purple_fill
    sheet['A2'].alignment = center_alignment
    sheet['A2'].font = bold_font
    sheet['A2'].border = thin_border

    # Preencher a coluna A com as entradas especificadas
    entries = [
        "ALA 12", "ALA 13", "ALA 14", "ALA 15", "ALA 16", "Total Bloco A", "REMIÇÃO 1", "REMIÇÃO 2",
        "ALA 1", "ALA 2", "ALA 3", "ALA 4", "ALA 5", "ALA 6", "ALA 7", "Total Bloco B",
        "Triagem 1", "Triagem 2", "Triagem 3", "ISOLAMENTO", "REMIÇÃO (ALA 12)", "Total Carceragem",
        "HGR (INTERNAÇÃO)", "TRAT. TOXICOLÓGICO", "PRISÃO DOMICILIAR", "SAÍDA TEMPORÁRIA",
        "SAÍDAS DO EXTERNO", "Total Externo", "TOTAL AO RECEBER", "ENTRADAS", "SAÍDAS", "",
        "TOTAL INTERNO", "TOTAL GERAL"
    ]
    current_row = 3
    for entry in entries:
        # Verifica se a entrada é vazia
        if entry == "":
            current_row += 1
            continue

        sheet[f'A{current_row}'] = entry
        sheet[f'A{current_row}'].alignment = center_alignment

        # Definir cor de fundo e bordas conforme a entrada
        if "Total" in entry or entry in ["SAÍDAS DO EXTERNO", "TOTAL AO RECEBER", "ENTRADAS", "SAÍDAS"]:
            sheet[f'A{current_row}'].fill = purple_fill if "Total" not in entry else yellow_fill
        else:
            sheet[f'A{current_row}'].fill = blue_fill

        sheet[f'A{current_row}'].border = thin_border
        current_row += 1

    # Mesclar células para os blocos e alas, e preencher com os dados de celas
    col_start = 3
    col_end = 4

    for block_key, block_name in [("A", "BLOCO A"), ("B", "BLOCO B")]:
        block_range_end = 'AA' if block_key == "B" else 'W'  # Ajuste para extensão do bloco B até a coluna AA
        sheet.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end + 8)
        sheet.merge_cells(start_row=1, start_column=14, end_row=1, end_column=27)
        sheet.cell(row=1, column=col_start, value=block_name).fill = gray_fill
        sheet.cell(row=1, column=col_start).alignment = center_alignment
        sheet.cell(row=1, column=col_start).font = bold_font

        for ala_key, ala_data in pamc_config[block_key]["alas"].items():
            # Verifica se a ala é "REMIÇÃO 01" ou "REMIÇÃO 02" e as ignora
            if ala_data["name"] in ["REMIÇÃO 01", "REMIÇÃO 02"]:
                continue

            # Mescla para o nome da Ala
            sheet.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
            sheet.cell(row=2, column=col_start, value=ala_data["name"]).alignment = center_alignment
            sheet.cell(row=2, column=col_start).font = bold_font

            # Preencher descrição
            sheet.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
            sheet.cell(row=3, column=col_start, value=ala_data["description"]).alignment = center_alignment

            # Preencher cabeçalho de celas e quantidade
            sheet.cell(row=4, column=col_start, value="CELA").alignment = center_alignment
            sheet.cell(row=4, column=col_end, value="QTD").alignment = center_alignment

            # Preencher celas
            start_cela_row = 5
            for i, cela in enumerate(ala_data["celas"], start=start_cela_row):
                sheet.cell(row=i, column=col_start, value=cela).fill = blue_fill
                sheet.cell(row=i, column=col_end).border = thin_border

            # Inserir REMIÇÃO 1 e REMIÇÃO 2 na Ala 01 (Bloco B)
            if block_key == "B" and ala_key == "01":
                sheet.cell(row=start_cela_row + len(ala_data["celas"]) + 2, column=col_start,
                           value="REM.1").fill = blue_fill
                sheet.cell(row=start_cela_row + len(ala_data["celas"]) + 3, column=col_start,
                           value="REM.2").fill = blue_fill

            # Ajustar colunas para próxima ala
            col_start += 2
            col_end += 2

        # Espaço entre Bloco A e Bloco B
        col_start += 1
        col_end += 1

    # Pintar de amarelo as células da linha 33 onde estiver "QTD"
    for col in ['D', 'F', 'H', 'J', 'L', 'O', 'Q', 'S', 'U', 'W', 'Y', 'AA']:
        sheet[f'{col}33'].fill = yellow_fill

    # Ajustar largura das colunas automaticamente
    sheet.column_dimensions['A'].width = 20  # Ajuste de exemplo; pode ser refinado
    sheet.column_dimensions['M'].width = 3  # Ajuste de exemplo; pode ser refinado

    return sheet


if __name__ == "__main__":
    workbook = Workbook()
    unit_name = "PAMC"
    control_sheet = generate_unit_control_sheet(workbook, unit_name)  # Agora recebe o 'sheet' diretamente

    # Transformar o caminho do arquivo em absoluto
    output_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Modelo_Controle.xlsx')
    workbook.save(output_file)
