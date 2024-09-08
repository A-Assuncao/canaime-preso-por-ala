from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import json
import os

# Define o diretório base relativo à localização do arquivo atual
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def load_units_config():
    # Carregar configurações do arquivo JSON
    config_path = os.path.join(BASE_DIR, 'units_config.json')
    with open(config_path, 'r', encoding='utf-8') as file:
        config_data = json.load(file)
    return config_data


def apply_borders(sheet, start_cell, end_cell):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in sheet[start_cell:end_cell]:
        for cell in row:
            cell.border = thin_border


def generate_unit_sei_sheet(workbook, unit_name):
    units_config = load_units_config()
    pamc_config = units_config[unit_name]["blocks"]

    # Cria a aba com o nome da unidade e "SEI"
    sheet = workbook.create_sheet(title=f"{unit_name} SEI")

    # Definir o tamanho das colunas de A até J para 8
    for col in range(1, 11):  # Colunas de A (1) até J (10)
        col_letter = sheet.cell(row=1, column=col).column_letter  # Obter a letra da coluna
        sheet.column_dimensions[col_letter].width = 8

    # Definir a fonte Arial 10
    arial_10_font = Font(name='Arial', size=10)

    # Aplicar a fonte de A1 até N69
    for row in range(1, 70):  # Linhas de 1 até 69
        for col in range(1, 15):  # Colunas de A (1) até N (14)
            cell = sheet.cell(row=row, column=col)
            cell.font = arial_10_font  # Aplicar fonte Arial 10

    # Definições de estilo
    gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')
    bold_font = Font(name='Arial', size=10, bold=True)

    # Configurações de Bloco A
    sheet.merge_cells('A1:J3')
    sheet['A1'] = "Bloco A"
    sheet['A1'].fill = gray_fill
    sheet['A1'].alignment = center_alignment
    sheet['A1'].font = bold_font

    # Ala 12
    sheet.merge_cells('A4:B4')
    sheet['A4'] = "Ala 12"
    sheet['A4'].alignment = center_alignment
    sheet['A4'].font = bold_font

    sheet['A5'] = "CELA"
    sheet['B5'] = "QTD"
    sheet['A5'].alignment = center_alignment
    sheet['B5'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["A"]["alas"]["12"]["celas"], start=6):
        sheet[f'A{i}'] = cela
        sheet[f'A{i}'].fill = blue_fill

    # Ala 13 ao lado de Ala 12
    sheet.merge_cells('C4:D4')
    sheet['C4'] = "Ala 13"
    sheet['C4'].alignment = center_alignment
    sheet['C4'].font = bold_font

    sheet['C5'] = "CELA"
    sheet['D5'] = "QTD"
    sheet['C5'].alignment = center_alignment
    sheet['D5'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["A"]["alas"]["13"]["celas"], start=6):
        sheet[f'C{i}'] = cela
        sheet[f'C{i}'].fill = blue_fill

    # Ala 14
    sheet.merge_cells('E4:F4')
    sheet['E4'] = "Ala 14"
    sheet['E4'].alignment = center_alignment
    sheet['E4'].font = bold_font

    sheet['E5'] = "CELA"
    sheet['F5'] = "QTD"
    sheet['E5'].alignment = center_alignment
    sheet['F5'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["A"]["alas"]["14"]["celas"], start=6):
        sheet[f'E{i}'] = cela
        sheet[f'E{i}'].fill = blue_fill

    # Ala 15
    sheet.merge_cells('G4:H4')
    sheet['G4'] = "Ala 15"
    sheet['G4'].alignment = center_alignment
    sheet['G4'].font = bold_font

    sheet['G5'] = "CELA"
    sheet['H5'] = "QTD"
    sheet['G5'].alignment = center_alignment
    sheet['H5'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["A"]["alas"]["15"]["celas"], start=6):
        sheet[f'G{i}'] = cela
        sheet[f'G{i}'].fill = blue_fill

    # Ala 16
    sheet.merge_cells('I4:J4')
    sheet['I4'] = "Ala 16"
    sheet['I4'].alignment = center_alignment
    sheet['I4'].font = bold_font

    sheet['I5'] = "CELA"
    sheet['J5'] = "QTD"
    sheet['I5'].alignment = center_alignment
    sheet['J5'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["A"]["alas"]["16"]["celas"], start=6):
        sheet[f'I{i}'] = cela
        sheet[f'I{i}'].fill = blue_fill

    # Configurar fundo amarelo na linha 34 para Bloco A, apenas em QTD
    for col in 'BDFHJ':
        sheet[f'{col}34'].fill = yellow_fill

    # Linha "TOTAL A"
    sheet["A35"] = "TOTAL A"
    sheet.merge_cells('B35:J35')
    sheet['A35'].alignment = center_alignment
    sheet['B35'].alignment = center_alignment

    # Configurações de Bloco B
    sheet.merge_cells('A37:N39')
    sheet['A37'] = "Bloco B"
    sheet['A37'].fill = gray_fill
    sheet['A37'].alignment = center_alignment
    sheet['A37'].font = bold_font

    # Ala 01
    sheet.merge_cells('A40:B40')
    sheet['A40'] = "Ala 01"
    sheet['A40'].alignment = center_alignment
    sheet['A40'].font = bold_font

    sheet['A41'] = "CELA"
    sheet['B41'] = "QTD"
    sheet['A41'].alignment = center_alignment
    sheet['B41'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["B"]["alas"]["01"]["celas"], start=42):
        sheet[f'A{i}'] = cela
        sheet[f'A{i}'].fill = blue_fill

    # Adiciona REM.1 e REM.2 depois de duas linhas em branco
    sheet['A48'] = ""  # Linha em branco
    sheet['A49'] = ""  # Linha em branco
    sheet['A50'] = "REM.1"
    sheet['A51'] = "REM.2"
    sheet['A50'].fill = blue_fill
    sheet['A51'].fill = blue_fill

    # Ala 02 ao lado de Ala 01
    sheet.merge_cells('C40:D40')
    sheet['C40'] = "Ala 02"
    sheet['C40'].alignment = center_alignment
    sheet['C40'].font = bold_font

    sheet['C41'] = "CELA"
    sheet['D41'] = "QTD"
    sheet['C41'].alignment = center_alignment
    sheet['D41'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["B"]["alas"]["02"]["celas"], start=42):
        sheet[f'C{i}'] = cela
        sheet[f'C{i}'].fill = blue_fill

    # Ala 03
    sheet.merge_cells('E40:F40')
    sheet['E40'] = "Ala 03"
    sheet['E40'].alignment = center_alignment
    sheet['E40'].font = bold_font

    sheet['E41'] = "CELA"
    sheet['F41'] = "QTD"
    sheet['E41'].alignment = center_alignment
    sheet['F41'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["B"]["alas"]["03"]["celas"], start=42):
        sheet[f'E{i}'] = cela
        sheet[f'E{i}'].fill = blue_fill

    # Ala 04
    sheet.merge_cells('G40:H40')
    sheet['G40'] = "Ala 04"
    sheet['G40'].alignment = center_alignment
    sheet['G40'].font = bold_font

    sheet['G41'] = "CELA"
    sheet['H41'] = "QTD"
    sheet['G41'].alignment = center_alignment
    sheet['H41'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["B"]["alas"]["04"]["celas"], start=42):
        sheet[f'G{i}'] = cela
        sheet[f'G{i}'].fill = blue_fill

    # Ala 05
    sheet.merge_cells('I40:J40')
    sheet['I40'] = "Ala 05"
    sheet['I40'].alignment = center_alignment
    sheet['I40'].font = bold_font

    sheet['I41'] = "CELA"
    sheet['J41'] = "QTD"
    sheet['I41'].alignment = center_alignment
    sheet['J41'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["B"]["alas"]["05"]["celas"], start=42):
        sheet[f'I{i}'] = cela
        sheet[f'I{i}'].fill = blue_fill

    # Ala 06
    sheet.merge_cells('K40:L40')
    sheet['K40'] = "Ala 06"
    sheet['K40'].alignment = center_alignment
    sheet['K40'].font = bold_font

    sheet['K41'] = "CELA"
    sheet['L41'] = "QTD"
    sheet['K41'].alignment = center_alignment
    sheet['L41'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["B"]["alas"]["06"]["celas"], start=42):
        sheet[f'K{i}'] = cela
        sheet[f'K{i}'].fill = blue_fill

    # Ala 07
    sheet.merge_cells('M40:N40')
    sheet['M40'] = "Ala 07"
    sheet['M40'].alignment = center_alignment
    sheet['M40'].font = bold_font

    sheet['M41'] = "CELA"
    sheet['N41'] = "QTD"
    sheet['M41'].alignment = center_alignment
    sheet['N41'].alignment = center_alignment

    for i, cela in enumerate(pamc_config["B"]["alas"]["07"]["celas"], start=42):
        sheet[f'M{i}'] = cela
        sheet[f'M{i}'].fill = blue_fill

    # Ajuste para o total do Bloco B na linha 66, fundo amarelo apenas em QTD
    for col in 'BDFHJLN':
        sheet[f'{col}66'].fill = yellow_fill

    # Total B, Triagem e Total Geral, com merge adequado e sem fundo amarelo
    sheet["A67"] = "TOTAL B"
    sheet.merge_cells('B67:N67')
    sheet['A67'].alignment = center_alignment

    sheet["A68"] = "TRIAGEM"
    sheet.merge_cells('B68:N68')
    sheet['A68'].alignment = center_alignment

    sheet["A69"] = "TOTAL GERAL"
    sheet.merge_cells('B69:N69')
    sheet['A69'].alignment = center_alignment

    # Aplicar bordas aos intervalos especificados
    apply_borders(sheet, 'A1', 'J35')
    apply_borders(sheet, 'A37', 'N69')

    return sheet


if __name__ == "__main__":
    workbook = Workbook()
    unit_name = "PAMC"
    control_sheet = generate_unit_sei_sheet(workbook, unit_name)  # Agora recebe o 'sheet' diretamente

    # Transformar o caminho do arquivo em absoluto
    output_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Modelo_SEI.xlsx')
    workbook.save(output_file)

