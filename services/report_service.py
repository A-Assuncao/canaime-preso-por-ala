from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string, get_column_letter
from config.excel_config_control import generate_unit_control_sheet
from config.excel_config_sei import generate_unit_sei_sheet
from tkinter import filedialog

from utils.logger import Logger

logger = Logger.get_logger()


def get_shift_name():
    """
    Retorna o nome do plantão com base na data atual.
    Ciclo de plantões: ALFA, BRAVO, CHARLIE, DELTA.
    """
    base_date = datetime(2024, 1, 1)  # Data de referência
    shifts = ['ALFA', 'BRAVO', 'CHARLIE', 'DELTA']
    current_date = datetime.now()
    delta_days = (current_date - base_date).days
    shift_name = shifts[delta_days % len(shifts)]
    return shift_name


def calculate_data(df):
    """
    Realiza cálculos nos dados da unidade, contando o número de presos por cela em cada ala e bloco.

    Parameters
    ----------
    df : DataFrame
        DataFrame contendo dados da unidade.

    Returns
    -------
    dict
        Dicionário com o número de presos por cela, ala e bloco.
    """
    # Agrupar pelos campos Bloco, Ala e Cela e contar quantos presos estão em cada combinação
    grouped_data = df.groupby(['Bloco', 'Ala', 'Cela']).size().reset_index(name='Quantidade')

    # Converter para um dicionário de fácil manipulação
    calculated_data = {}
    for _, row in grouped_data.iterrows():
        bloco = row['Bloco']
        ala = row['Ala']
        cela = row['Cela']
        quantidade = row['Quantidade']

        # Estrutura o dicionário conforme a hierarquia Bloco -> Ala -> Cela
        if bloco not in calculated_data:
            calculated_data[bloco] = {}
        if ala not in calculated_data[bloco]:
            calculated_data[bloco][ala] = {}

        calculated_data[bloco][ala][cela] = quantidade

    return calculated_data


def fill_control_sheet(ws, data):
    """
    Preenche a aba "Controle" com os dados calculados para cada cela e faz o somatório na linha 33.
    Alinha as células no centro e calcula o somatório das celas de "REMIÇÃO 01" e "REMIÇÃO 02".

    Parameters
    ----------
    ws : Worksheet
        A aba de planilha do workbook a ser preenchida.
    data : dict
        Dicionário com os dados calculados, contendo a quantidade de presos por cela
        (aninhado por Bloco -> Ala -> Cela).
    """
    center_alignment = Alignment(horizontal='center', vertical='center')  # Alinhamento centralizado
    # Mapeamento das colunas por bloco e alas
    bloco_a_cols = ['D', 'F', 'H', 'J', 'L']  # Alas 12, 13, 14, 15, 16
    bloco_b_cols = ['O', 'Q', 'S', 'U', 'W', 'Y', 'AA']  # Alas 01, 02, 03, 04, 05, 06, 07
    row_start = 5
    row_end = 32
    total_row = 33

    # Preencher Bloco A
    bloco_a_alas = ['12', '13', '14', '15', '16']
    for col, ala in zip(bloco_a_cols, bloco_a_alas):
        for row in range(row_start, row_end + 1):
            cela_value = ws[f'{get_column_letter(column_index_from_string(col) - 1)}{row}'].value
            if cela_value is None:
                continue

            cela = str(cela_value)
            if 'A' in data and ala in data['A'] and cela in data['A'][ala]:
                quantidade_presos = data['A'][ala][cela]
                ws[f'{col}{row}'] = quantidade_presos
                ws[f'{col}{row}'].alignment = center_alignment  # Centraliza o valor
            else:
                ws[f'{col}{row}'] = 0
                ws[f'{col}{row}'].alignment = center_alignment  # Centraliza o valor

        # Somatório da coluna na linha 33
        ws[f'{col}{total_row}'] = f'=SUM({col}{row_start}:{col}{row_end})'
        ws[f'{col}{total_row}'].alignment = center_alignment  # Centraliza o somatório

    # Preencher Bloco B
    bloco_b_alas = ['01', '02', '03', '04', '05', '06', '07']
    for col, ala in zip(bloco_b_cols, bloco_b_alas):
        for row in range(row_start, row_end + 1):
            cela_value = ws[f'{get_column_letter(column_index_from_string(col) - 1)}{row}'].value
            if cela_value is None:
                continue

            cela = str(cela_value)
            if 'B' in data and ala in data['B'] and cela in data['B'][ala]:
                quantidade_presos = data['B'][ala][cela]
                ws[f'{col}{row}'] = quantidade_presos
                ws[f'{col}{row}'].alignment = center_alignment  # Centraliza o valor
            else:
                ws[f'{col}{row}'] = 0
                ws[f'{col}{row}'].alignment = center_alignment  # Centraliza o valor

        # Somatório da coluna na linha 33
        ws[f'{col}{total_row}'] = f'=SUM({col}{row_start}:{col}{row_end})'
        ws[f'{col}{total_row}'].alignment = center_alignment  # Centraliza o somatório

    # Somatório de todas as celas de REMIÇÃO 01 e REMIÇÃO 02
    rem_01_total = 0
    rem_02_total = 0

    if 'B' in data:
        # Somar todas as celas dentro de REMIÇÃO 01
        if 'REMIÇÃO01' in data['B']:
            rem_01_total = data['B']['REMIÇÃO01'].get('1', 0)

        # Somar todas as celas dentro de REMIÇÃO 02
        if 'REMIÇÃO02' in data['B']:
            rem_02_total = data['B']['REMIÇÃO02'].get('2', 0)

    # Preencher o somatório total em O13 para REMIÇÃO 01 e O14 para REMIÇÃO 02
    ws['O13'] = rem_01_total
    ws['O13'].alignment = center_alignment  # Centraliza o valor

    ws['O14'] = rem_02_total
    ws['O14'].alignment = center_alignment  # Centraliza o valor

    # Preenchendo fórmulas e instruções nas células de B3 até B36 com alinhamento centralizado
    ws['B3'] = '=D33'
    ws['B4'] = '=F33'
    ws['B5'] = '=H33'
    ws['B6'] = '=J33'
    ws['B7'] = '=L33'
    ws['B8'] = '=SUM(B3:B7)'
    ws['B9'] = '=O13'
    ws['B10'] = '=O14'
    ws['B11'] = '=O33'
    ws['B12'] = '=Q33'
    ws['B13'] = '=S33'
    ws['B14'] = '=U33'
    ws['B15'] = '=W33'
    ws['B16'] = '=Y33'
    ws['B17'] = '=AA33'
    ws['B18'] = '=SUM(B9:B17)'
    ws['B24'] = '=SUM(B19:B23)'
    ws['B30'] = '=SUM(B25:B28)'
    ws['B35'] = '=SUM(B8,B18,B24)'
    ws['B36'] = '=SUM(B30,B35)'

    # Somatório da ala TRIAGEM
    triagem_total = 0
    if 'Carceragem' in data and 'TRIAGEM' in data['Carceragem']:
        for cela, quantidade in data['Carceragem']['TRIAGEM'].items():
            triagem_total += quantidade

    # Preencher o somatório total em B19 para TRIAGEM
    ws['B19'] = triagem_total
    ws['B19'].alignment = Alignment(horizontal='center', vertical='center')  # Centraliza o valor

    # Somatório de todas as celas de Externo (HGR, TRATOX, PRIS/DOM)
    externo_hgr_total = 0
    externo_tratox_total = 0
    externo_prisdom_total = 0

    if 'Externo' in data:
        # Somar todas as celas dentro de HGR
        if 'HGR' in data['Externo']:
            for cela, qtd in data['Externo']['HGR'].items():
                externo_hgr_total += qtd
            ws['B25'] = externo_hgr_total

        # Somar todas as celas dentro de TRATOX
        if 'TRATOX' in data['Externo']:
            for cela, qtd in data['Externo']['TRATOX'].items():
                externo_tratox_total += qtd
            ws['B26'] = externo_tratox_total

        # Somar todas as celas dentro de PRIS/DOM
        if 'PRIS/DOM' in data['Externo']:
            for cela, qtd in data['Externo']['PRIS/DOM'].items():
                externo_prisdom_total += qtd
            ws['B27'] = externo_prisdom_total

    # Alinhamento centralizado das células de B3 até B36
    for row in range(3, 37):
        ws[f'B{row}'].alignment = center_alignment


def fill_sei_sheet(ws_sei, ws_control):
    """
    Preenche a aba "SEI" com os dados da aba "Controle".

    Parameters
    ----------
    ws_sei : Worksheet
        A aba "SEI" do workbook a ser preenchida.
    ws_control : Worksheet
        A aba "Controle" do workbook de onde os dados serão extraídos.
    """
    # Correspondência das células entre a aba "Controle" e a aba "SEI"
    cell_mapping = [
        ('B6', 'D5', 'D6', 'F5', 'F6', 'H5', 'H6', 'J5', 'J6', 'L5'),
        ('B7', 'D6', 'D7', 'F6', 'F7', 'H6', 'H7', 'J6', 'J7', 'L6'),
        ('B8', 'D7', 'D8', 'F7', 'F8', 'H7', 'H8', 'J7', 'J8', 'L7'),
        ('B9', 'D8', 'D9', 'F8', 'F9', 'H8', 'H9', 'J8', 'J9', 'L8'),
        ('B10', 'D9', 'D10', 'F9', 'F10', 'H9', 'H10', 'J9', 'J10', 'L9'),
        ('B11', 'D10', 'D11', 'F10', 'F11', 'H10', 'H11', 'J10', 'J11', 'L10'),
        ('B12', 'D11', 'D12', 'F11', 'F12', 'H11', 'H12', 'J11', 'J12', 'L11'),
        ('B13', 'D12', 'D13', 'F12', 'F13', 'H12', 'H13', 'J12', 'J13', 'L12'),
        ('B14', 'D13', 'D14', 'F13', 'F14', 'H13', 'H14', 'J13', 'J14', 'L13'),
        ('B15', 'D14', 'D15', 'F14', 'F15', 'H14', 'H15', 'J14', 'J15', 'L14'),
        ('B16', 'D15', 'D16', 'F15', 'F16', 'H15', 'H16', 'J15', 'J16', 'L15'),
        ('B17', 'D16', 'D17', 'F16', 'F17', 'H16', 'H17', 'J16', 'J17', 'L16'),
        ('B18', 'D17', 'D18', 'F17', 'F18', 'H17', 'H18', 'J17', 'J18', 'L17'),
        ('B19', 'D18', 'D19', 'F18', 'F19', 'H18', 'H19', 'J18', 'J19', 'L18'),
        ('B20', 'D19', 'D20', 'F19', 'F20', 'H19', 'H20', 'J19', 'J20', 'L19'),
        ('B21', 'D20', 'D21', 'F20', 'F21', 'H20', 'H21', 'J20', 'J21', 'L20'),
        ('B22', 'D21', 'D22', 'F21', 'F22', 'H21', 'H22', 'J21', 'J22', 'L21'),
        ('B23', 'D22', 'D23', 'F22', 'F23', 'H22', 'H23', 'J22', 'J23', 'L22'),
        ('B24', 'D23', 'D24', 'F23', 'F24', 'H23', 'H24', 'J23', 'J24', 'L23'),
        ('B25', 'D24', 'D25', 'F24', 'F25', 'H24', 'H25', 'J24', 'J25', 'L24'),
        ('B26', 'D25', 'D26', 'F25', 'F26', 'H25', 'H26', 'J25', 'J26', 'L25'),
        ('B27', 'D26', 'D27', 'F26', 'F27', 'H26', 'H27', 'J26', 'J27', 'L26'),
        ('B28', 'D27', 'D28', 'F27', 'F28', 'H27', 'H28', 'J27', 'J28', 'L27'),
        ('B29', 'D28', 'D29', 'F28', 'F29', 'H28', 'H29', 'J28', 'J29', 'L28'),
        ('B30', 'D29', 'D30', 'F29', 'F30', 'H29', 'H30', 'J29', 'J30', 'L29'),
        ('B31', 'D30', 'D31', 'F30', 'F31', 'H30', 'H31', 'J30', 'J31', 'L30'),
        ('B32', 'D31', 'D32', 'F31', 'F32', 'H31', 'H32', 'J31', 'J32', 'L31'),
        ('B33', 'D32', 'D33', 'F32', 'F33', 'H32', 'H33', 'J32', 'J33', 'L32'),
        ('B34', 'D33', 'D34', 'F33', 'F34', 'H33', 'H34', 'J33', 'J34', 'L33')
    ]

    # Preenchendo os valores na aba SEI com base nos valores da aba Controle
    for sei_cells in cell_mapping:
        ws_sei[sei_cells[0]] = ws_control[sei_cells[1]].value
        ws_sei[sei_cells[2]] = ws_control[sei_cells[3]].value
        ws_sei[sei_cells[4]] = ws_control[sei_cells[5]].value
        ws_sei[sei_cells[6]] = ws_control[sei_cells[7]].value
        ws_sei[sei_cells[8]] = ws_control[sei_cells[9]].value

    # Inserindo fórmulas de soma na linha 34
    ws_sei['B34'] = '=SUM(B6:B33)'
    ws_sei['D34'] = '=SUM(D6:D33)'
    ws_sei['F34'] = '=SUM(F6:F33)'
    ws_sei['H34'] = '=SUM(H6:H33)'
    ws_sei['J34'] = '=SUM(J6:J33)'

    ws_sei['B35'] = '=SUM(B34:J34)'  # Total A

    # Centralizando o conteúdo das células da aba SEI
    for row in ws_sei.iter_rows(min_row=6, max_row=34, min_col=2, max_col=10):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Mapeamento de células para a segunda parte (B42 até N66)
    cell_mapping_part_2 = [
        ('B42', 'O5'), ('D42', 'Q5'), ('F42', 'S5'), ('H42', 'U5'), ('J42', 'W5'), ('L42', 'Y5'), ('N42', 'AA5'),
        ('B43', 'O6'), ('D43', 'Q6'), ('F43', 'S6'), ('H43', 'U6'), ('J43', 'W6'), ('L43', 'Y6'), ('N43', 'AA6'),
        ('B44', 'O7'), ('D44', 'Q7'), ('F44', 'S7'), ('H44', 'U7'), ('J44', 'W7'), ('L44', 'Y7'), ('N44', 'AA7'),
        ('B45', 'O8'), ('D45', 'Q8'), ('F45', 'S8'), ('H45', 'U8'), ('J45', 'W8'), ('L45', 'Y8'), ('N45', 'AA8'),
        ('B46', 'O9'), ('D46', 'Q9'), ('F46', 'S9'), ('H46', 'U9'), ('J46', 'W9'), ('L46', 'Y9'), ('N46', 'AA9'),
        ('B47', 'O10'), ('D47', 'Q10'), ('F47', 'S10'), ('H47', 'U10'), ('J47', 'W10'), ('L47', 'Y10'), ('N47', 'AA10'),
        ('B48', 'O11'), ('D48', 'Q11'), ('F48', 'S11'), ('H48', 'U11'), ('J48', 'W11'), ('L48', 'Y11'), ('N48', 'AA11'),
        ('B49', 'O12'), ('D49', 'Q12'), ('F49', 'S12'), ('H49', 'U12'), ('J49', 'W12'), ('L49', 'Y12'), ('N49', 'AA12'),
        ('B50', 'O13'), ('D50', 'Q13'), ('F50', 'S13'), ('H50', 'U13'), ('J50', 'W13'), ('L50', 'Y13'), ('N50', 'AA13'),
        ('B51', 'O14'), ('D51', 'Q14'), ('F51', 'S14'), ('H51', 'U14'), ('J51', 'W14'), ('L51', 'Y14'), ('N51', 'AA14'),
        ('B52', 'O15'), ('D52', 'Q15'), ('F52', 'S15'), ('H52', 'U15'), ('J52', 'W15'), ('L52', 'Y15'), ('N52', 'AA15'),
        ('B53', 'O16'), ('D53', 'Q16'), ('F53', 'S16'), ('H53', 'U16'), ('J53', 'W16'), ('L53', 'Y16'), ('N53', 'AA16'),
        ('B54', 'O17'), ('D54', 'Q17'), ('F54', 'S17'), ('H54', 'U17'), ('J54', 'W17'), ('L54', 'Y17'), ('N54', 'AA17'),
        ('B55', 'O18'), ('D55', 'Q18'), ('F55', 'S18'), ('H55', 'U18'), ('J55', 'W18'), ('L55', 'Y18'), ('N55', 'AA18'),
        ('B56', 'O19'), ('D56', 'Q19'), ('F56', 'S19'), ('H56', 'U19'), ('J56', 'W19'), ('L56', 'Y19'), ('N56', 'AA19'),
        ('B57', 'O20'), ('D57', 'Q20'), ('F57', 'S20'), ('H57', 'U20'), ('J57', 'W20'), ('L57', 'Y20'), ('N57', 'AA20'),
        ('B58', 'O21'), ('D58', 'Q21'), ('F58', 'S21'), ('H58', 'U21'), ('J58', 'W21'), ('L58', 'Y21'), ('N58', 'AA21'),
        ('B59', 'O22'), ('D59', 'Q22'), ('F59', 'S22'), ('H59', 'U22'), ('J59', 'W22'), ('L59', 'Y22'), ('N59', 'AA22'),
        ('B60', 'O23'), ('D60', 'Q23'), ('F60', 'S23'), ('H60', 'U23'), ('J60', 'W23'), ('L60', 'Y23'), ('N60', 'AA23'),
        ('B61', 'O24'), ('D61', 'Q24'), ('F61', 'S24'), ('H61', 'U24'), ('J61', 'W24'), ('L61', 'Y24'), ('N61', 'AA24'),
        ('B62', 'O25'), ('D62', 'Q25'), ('F62', 'S25'), ('H62', 'U25'), ('J62', 'W25'), ('L62', 'Y25'), ('N62', 'AA25'),
        ('B63', 'O26'), ('D63', 'Q26'), ('F63', 'S26'), ('H63', 'U26'), ('J63', 'W26'), ('L63', 'Y26'), ('N63', 'AA26'),
        ('B64', 'O27'), ('D64', 'Q27'), ('F64', 'S27'), ('H64', 'U27'), ('J64', 'W27'), ('L64', 'Y27'), ('N64', 'AA27'),
        ('B65', 'O28'), ('D65', 'Q28'), ('F65', 'S28'), ('H65', 'U28'), ('J65', 'W28'), ('L65', 'Y28'), ('N65', 'AA28')
    ]

    # Preenchendo os valores na aba SEI com base nos valores da aba Controle
    for sei_cell, control_cell in cell_mapping_part_2:
        ws_sei[sei_cell] = ws_control[control_cell].value

    # Inserindo fórmulas de soma na linha 66
    ws_sei['B66'] = '=SUM(B42:B65)'
    ws_sei['D66'] = '=SUM(D42:D65)'
    ws_sei['F66'] = '=SUM(F42:F65)'
    ws_sei['H66'] = '=SUM(H42:H65)'
    ws_sei['J66'] = '=SUM(J42:J65)'
    ws_sei['L66'] = '=SUM(L42:L65)'
    ws_sei['N66'] = '=SUM(N42:N65)'

    name_ws_control_sheet = ws_control.title
    triagem = f"=SUM('{name_ws_control_sheet}'!B19:B21)"

    ws_sei['B67'] = '=SUM(B66:N66)'  # Total B
    ws_sei['B68'] = triagem  # Triagem
    ws_sei['B69'] = '=SUM(B67:B68)'  # total geral

    # Centralizando o conteúdo das células da aba SEI
    for row in ws_sei.iter_rows(min_row=42, max_row=69, min_col=2, max_col=14):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


def create_excel_report(data):
    """
    Cria um relatório Excel para os dados fornecidos e salva no caminho especificado.

    Parameters
    ----------
    data : dict
        Dicionário contendo os dados das unidades.

    """
    logger.info("Iniciando criação do relatório Excel.")

    # Criar um novo workbook
    wb = Workbook()

    try:
        for unit_name, df_list in data.items():
            logger.debug(f"Processando unidade: {unit_name}")
            # Converter os dados da unidade em um DataFrame
            if df_list:
                df = pd.DataFrame(df_list)

                # Gerar as abas de controle e SEI
                control_ws = generate_unit_control_sheet(wb, unit_name)
                sei_ws = generate_unit_sei_sheet(wb, unit_name)

                # Realizar cálculos nos dados
                calculated_data = calculate_data(df)
                logger.debug(f"Dados calculados: {calculated_data}")

                # Preencher as abas de controle e SEI
                fill_control_sheet(control_ws, calculated_data)
                fill_sei_sheet(sei_ws, control_ws)

        # Remover a aba padrão "Sheet"
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

            # Gerar nome do arquivo baseado no plantão e data
            shift_name = get_shift_name()
            current_date = datetime.now().strftime('%d-%m-%Y')
            default_filename = f"Contagem-{shift_name}-{current_date}.xlsx"

            # Abrir a caixa de diálogo para o usuário escolher onde salvar o arquivo
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_filename,
                                                     filetypes=[("Excel files", "*.xlsx")])

            if file_path:
                wb.save(file_path)
                logger.info(f"Relatório salvo com sucesso em: {file_path}")
            else:
                logger.warning("Salvamento cancelado pelo usuário.")
    except Exception as e:
        Logger.capture_error(e)
        logger.error(f"Erro ao criar o relatório Excel: {str(e)}")
